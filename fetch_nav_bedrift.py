#!/usr/bin/env python3
"""Parser NAVs bedriftsundersøkelse (Excel-fil) og injiserer som inline
NAV_BEDRIFT i assets/data.js.

NAV publiserer årlig en undersøkelse av virksomheters arbeidskraftsbehov.
2026-utgaven har fylkesnedbrytning som lar oss aggregere til Nord-Norge.

Kilde: https://bibliotek.nav.no/record/6225
       Navs bedriftsundersøkelse 2026 — Figurer og tabeller

Tabeller vi henter:
  Tabell V1   — Estimert mangel på arbeidskraft per næring × fylke (2026)
  Tabell 1/5  — Mangel + stramhetsindikator per næring (nasjonalt)
  Figur 6     — Sysselsettingsbarometer per fylke (2025 og 2026)

Datastrukturen som injiseres:
  NAV_BEDRIFT = {
    retrieved_at, source, source_url, year,
    nasjonalt: { mangel_total, naeringer: [...] },
    landsdel:  { mangel_total, mangel_andel_av_nasjonalt,
                 sysselsetting_barometer_2026_vektet,
                 naeringer: [{ navn, mangel_nn, mangel_nat, andel_pct }, ...] },
    fylker:    { Nordland: {mangel_total, sysselsetting_barometer_2025/2026,
                            naeringer: [...]}, Troms: {...}, Finnmark: {...} }
  }
"""
import json
import re
import sys
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HERE = Path(__file__).parent
HTML = HERE / "assets" / "data.js"
DEFAULT_XLSX = Path.home() / "Downloads" / "Navs bedriftsundersøkelse 2026 Figurer og tabeller.xlsx"

YEAR = 2026
SOURCE_NAME = "NAVs bedriftsundersøkelse 2026"
SOURCE_URL = "https://bibliotek.nav.no/record/6225"
NN_FYLKER = ("Nordland", "Troms", "Finnmark")


def load_workbook(path: Path):
    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def parse_v1(wb):
    """Tabell V1: Estimert mangel per næring × fylke.
    Returnerer (naering_list, headers_list, matrix, totals).
    Filtrerer ut underkategorier (rader som starter med '-') så vi unngår
    dobbeltelling — disse er underkategorier av 'Industrien samlet'."""
    ws = wb["Tabell V1 - 26"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    headers = rows[2]
    fylker = [h for h in headers[1:16]]
    naeringer = []
    matrix = []
    totals = None
    for r in rows[3:]:
        if not r[0]:
            continue
        name = str(r[0]).strip().rstrip("'")
        if name == "I alt":
            totals = [r[j+1] or 0 for j in range(15)]
            continue
        # Hopp over underkategorier av Industrien (rader som starter med '-')
        if name.startswith("-"):
            continue
        vals = [(r[j+1] or 0) for j in range(15)]
        naeringer.append(name)
        matrix.append(vals)
    return naeringer, fylker, matrix, totals


def parse_figur6(wb):
    """Figur 6: Sysselsettingsbarometer (nettoandel) per fylke, 2025 og 2026."""
    ws = wb["Figur 6 - 26"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, "", "Figur 6. Navs sysselsettingsbarometer for 2025 og 2026, etter fylke"):
            continue
        name = str(row[0]).strip()
        if name in ("",):
            continue
        if name in (None,):
            continue
        if row[1] is None:
            continue
        try:
            v25 = float(row[1])
            v26 = float(row[2]) if row[2] is not None else None
            out[name] = {"2025": round(v25, 2), "2026": round(v26, 2) if v26 is not None else None}
        except (TypeError, ValueError):
            continue
    return out


def parse_tabell5(wb):
    """Tabell 5: Mangel + stramhetsindikator per næring (nasjonalt)."""
    ws = wb["Tabell 5 - 26"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    out = {}
    for r in rows[2:]:
        if not r[0]:
            continue
        name = str(r[0]).strip().rstrip("'")
        try:
            mangel = int(r[1]) if r[1] is not None else None
            ci_lo = int(r[2]) if r[2] is not None else None
            ci_hi = int(r[3]) if r[3] is not None else None
            stramhet = round(float(r[4]), 2) if r[4] is not None else None
            alvorlige = round(float(r[5]), 3) if r[5] is not None else None
        except (TypeError, ValueError):
            continue
        out[name] = {
            "mangel": mangel,
            "ci": [ci_lo, ci_hi] if ci_lo is not None and ci_hi is not None else None,
            "stramhet": stramhet,
            "alvorlige_andel": alvorlige,
        }
    return out


def build_payload(xlsx_path: Path):
    wb = load_workbook(xlsx_path)

    # 1. V1: mangel-matrise
    naeringer, fylker, matrix, totals = parse_v1(wb)
    fy_idx = {f: i for i, f in enumerate(fylker)}

    # Nasjonale totaler per næring (sum over alle fylker i V1)
    nasjonalt_naeringer = []
    nat_total = 0
    nn_total = 0
    for i, naering in enumerate(naeringer):
        nat = sum(matrix[i])
        nn = sum(matrix[i][fy_idx[f]] for f in NN_FYLKER)
        nasjonalt_naeringer.append({"navn": naering, "mangel": nat})
        nat_total += nat
        nn_total += nn

    # Landsdel: aggregert + topp-næringer sortert
    landsdel_naeringer = []
    for i, naering in enumerate(naeringer):
        nat = sum(matrix[i])
        nn = sum(matrix[i][fy_idx[f]] for f in NN_FYLKER)
        if nn <= 0 and nat <= 0:
            continue
        andel = round(nn / nat * 100, 1) if nat else None
        landsdel_naeringer.append({
            "navn": naering,
            "mangel_nn": nn,
            "mangel_nat": nat,
            "andel_pct": andel,
        })
    landsdel_naeringer.sort(key=lambda x: -x["mangel_nn"])

    # Per fylke
    fylker_data = {}
    for f in NN_FYLKER:
        col = fy_idx[f]
        f_naeringer = []
        f_total = 0
        for i, naering in enumerate(naeringer):
            v = matrix[i][col]
            if v <= 0:
                continue
            f_naeringer.append({"navn": naering, "mangel": v})
            f_total += v
        f_naeringer.sort(key=lambda x: -x["mangel"])
        fylker_data[f] = {
            "mangel_total": f_total,
            "naeringer": f_naeringer,
        }

    # 2. Figur 6: sysselsettingsbarometer per fylke
    barometer = parse_figur6(wb)
    for f in NN_FYLKER:
        if f in barometer:
            fylker_data[f]["sysselsetting_barometer"] = barometer[f]
        elif f"{f} " in barometer:  # noen ark har trailing space
            fylker_data[f]["sysselsetting_barometer"] = barometer[f"{f} "]

    # Vektet snitt for landsdel (vekt med fylkets mangel-total)
    sb_w_sum = 0
    sb_total_w = 0
    for f in NN_FYLKER:
        sb = fylker_data[f].get("sysselsetting_barometer")
        if sb and sb.get("2026") is not None:
            w = fylker_data[f]["mangel_total"]
            sb_w_sum += sb["2026"] * w
            sb_total_w += w
    landsdel_sb_2026 = round(sb_w_sum / sb_total_w, 1) if sb_total_w else None

    # 3. Tabell 5: nasjonale tall (totalt mangel + stramhet)
    t5 = parse_tabell5(wb)
    nat_overall = t5.get("I alt", {}).get("mangel")
    nat_stramhet = t5.get("I alt", {}).get("stramhet")
    nat_ci = t5.get("I alt", {}).get("ci")

    # Befolkningsandel som referanse (Nord-Norge ca. 8,8 % av Norge)
    NN_POP_SHARE = 8.8

    payload = {
        "retrieved_at": datetime.date.today().isoformat(),
        "source": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "year": YEAR,
        "nasjonalt": {
            "mangel_total": nat_overall,
            "mangel_ci": nat_ci,
            "stramhet": nat_stramhet,
            "sysselsetting_barometer_2026": barometer.get("I alt", {}).get("2026"),
            "naeringer_top": sorted(nasjonalt_naeringer, key=lambda x: -x["mangel"])[:10],
        },
        "landsdel": {
            "mangel_total": nn_total,
            "mangel_andel_av_nasjonalt_pct": round(nn_total / nat_total * 100, 1) if nat_total else None,
            "befolkning_andel_pct": NN_POP_SHARE,
            "sysselsetting_barometer_2026_vektet": landsdel_sb_2026,
            "naeringer": landsdel_naeringer,
        },
        "fylker": fylker_data,
    }
    return payload


def inject(payload, html_path: Path):
    html = html_path.read_text(encoding="utf-8")
    js_lit = "const NAV_BEDRIFT = " + json.dumps(payload, ensure_ascii=False, indent=0).replace("\n", "") + ";"
    marker_start = "/* NAV_BEDRIFT BEGIN */"
    marker_end = "/* NAV_BEDRIFT END */"
    block = f"{marker_start}\n{js_lit}\n{marker_end}"
    if marker_start in html:
        pat = re.compile(r"/\* NAV_BEDRIFT BEGIN \*/.*?/\* NAV_BEDRIFT END \*/", re.DOTALL)
        new_html = pat.sub(block, html)
    else:
        # Plasser etter siste relaterte data-blokk
        for anchor in ("/* LEVEKAR_DATA END */", "/* ARBEID_DATA END */", "/* RENTE_DATA END */"):
            if anchor in html:
                new_html = html.replace(anchor, anchor + "\n" + block, 1)
                break
        else:
            print("ERROR: ingen kjent anchor — legg block manuelt", file=sys.stderr)
            sys.exit(1)
    html_path.write_text(new_html, encoding="utf-8")
    print(f"Injected NAV_BEDRIFT ({len(js_lit)} chars) into {html_path.name}")


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    print(f"Reading {xlsx}")
    payload = build_payload(xlsx)
    nn = payload["landsdel"]
    print(f"\nNord-Norge: {nn['mangel_total']} mangler (av {payload['nasjonalt']['mangel_total']} nasjonalt)")
    print(f"  Andel av nasjonalt: {nn['mangel_andel_av_nasjonalt_pct']} %  (befolkningsandel ~{nn['befolkning_andel_pct']} %)")
    print(f"  Vektet sysselsetting-barometer 2026: {nn['sysselsetting_barometer_2026_vektet']}")
    print(f"\nTopp 6 mangel-næringer i Nord-Norge:")
    for n in nn["naeringer"][:6]:
        print(f"  {n['navn'][:40]:42s} NN={n['mangel_nn']:5d}  ({n['andel_pct']:.0f} % av nasjonalt)")
    print(f"\nPer fylke:")
    for f, fd in payload["fylker"].items():
        sb = fd.get("sysselsetting_barometer", {})
        print(f"  {f:12s}  total={fd['mangel_total']:5d}  barometer 2025={sb.get('2025')}  2026={sb.get('2026')}")
    inject(payload, HTML)


if __name__ == "__main__":
    main()
